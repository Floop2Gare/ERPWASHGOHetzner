"""
Script pour exporter toutes les prestations Google Calendar vers un fichier Excel
Période: 1er janvier 2025 au 8 janvier 2026
"""
import os
import sys
from datetime import datetime
from pathlib import Path

# Configurer l'encodage UTF-8 pour la console Windows
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Ajouter le répertoire backend/app au PYTHONPATH
backend_dir = Path(__file__).parent
app_dir = backend_dir / "app"
sys.path.insert(0, str(backend_dir))
sys.path.insert(0, str(app_dir))

# Définir les chemins des fichiers credentials pour l'environnement local
credentials_adrien_path = backend_dir / 'credentials_adrien.json'
credentials_clement_path = backend_dir / 'credentials_clement.json'

# Essayer de charger les credentials depuis les fichiers si disponibles
# Sinon, le service utilisera les variables d'environnement
if credentials_adrien_path.exists():
    try:
        # Essayer de charger et définir comme variable d'environnement
        with open(credentials_adrien_path, 'r', encoding='utf-8') as f:
            import json
            creds = json.load(f)
            os.environ['GOOGLE_SA_ADRIEN_JSON'] = json.dumps(creds)
    except Exception as e:
        print(f"[AVERTISSEMENT] Impossible de charger credentials_adrien.json: {e}")
        print("   Le script utilisera les variables d'environnement si disponibles.")

if credentials_clement_path.exists():
    try:
        with open(credentials_clement_path, 'r', encoding='utf-8') as f:
            import json
            creds = json.load(f)
            os.environ['GOOGLE_SA_CLEMENT_JSON'] = json.dumps(creds)
    except Exception as e:
        print(f"[AVERTISSEMENT] Impossible de charger credentials_clement.json: {e}")
        print("   Le script utilisera les variables d'environnement si disponibles.")

# Définir aussi les chemins de fichiers pour le service
os.environ['GOOGLE_SA_ADRIEN_FILE'] = str(credentials_adrien_path)
os.environ['GOOGLE_SA_CLEMENT_FILE'] = str(credentials_clement_path)

from app.services.google_calendar import GoogleCalendarService
from googleapiclient.errors import HttpError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

def parse_datetime(dt_str: str) -> datetime:
    """Parse une date ISO en datetime Python"""
    try:
        if not dt_str:
            return None
        
        # Format avec heure: 2025-01-01T10:00:00+01:00 ou 2025-01-01T10:00:00Z
        if 'T' in dt_str:
            # Enlever le fuseau horaire pour simplifier
            if '+' in dt_str:
                dt_str = dt_str.split('+')[0]
            elif dt_str.endswith('Z'):
                dt_str = dt_str[:-1]
            return datetime.fromisoformat(dt_str)
        # Format date seule: 2025-01-01
        else:
            return datetime.fromisoformat(dt_str)
    except Exception as e:
        print(f"Erreur lors du parsing de la date '{dt_str}': {e}")
        return None

def format_datetime_fr(dt: datetime) -> str:
    """Formate une datetime en format français"""
    if dt is None:
        return ""
    return dt.strftime("%d/%m/%Y %H:%M")

def format_date_fr(dt: datetime) -> str:
    """Formate une date en format français"""
    if dt is None:
        return ""
    return dt.strftime("%d/%m/%Y")

def extract_duration(start_str: str, end_str: str) -> str:
    """Calcule la durée d'un événement"""
    start = parse_datetime(start_str) if start_str else None
    end = parse_datetime(end_str) if end_str else None
    
    if start and end:
        duration = end - start
        hours = duration.total_seconds() / 3600
        if hours < 1:
            minutes = int(duration.total_seconds() / 60)
            return f"{minutes} min"
        else:
            return f"{hours:.1f} h"
    return ""

def clean_description(desc: str) -> str:
    """Nettoie et formate la description"""
    if not desc:
        return ""
    # Retirer les caractères de saut de ligne multiples
    desc = desc.replace('\r\n', ' ').replace('\n', ' ')
    # Retirer les espaces multiples
    return ' '.join(desc.split())

def extract_phone_number(text: str) -> str:
    """Extrait le numéro de téléphone depuis un texte"""
    if not text:
        return ""
    
    # Patterns pour numéros français
    patterns = [
        r'(\+33\s?[1-9](?:[\s.-]?\d{2}){4})',  # +33 X XX XX XX XX
        r'(0[1-9](?:[\s.-]?\d{2}){4})',  # 0X XX XX XX XX
        r'(\d{2}[\s.-]?\d{2}[\s.-]?\d{2}[\s.-]?\d{2}[\s.-]?\d{2})',  # XX XX XX XX XX
        r'(tel[:\s]+([0-9\s\.\-]+))',  # tel: 06 12 34 56 78
        r'(téléphone[:\s]+([0-9\s\.\-]+))',  # téléphone: 06 12 34 56 78
        r'(phone[:\s]+([0-9\s\.\-]+))',  # phone: 06 12 34 56 78
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            # Prendre le premier match
            phone = matches[0]
            if isinstance(phone, tuple):
                phone = phone[1] if phone[1] else phone[0]
            # Nettoyer le numéro
            phone = re.sub(r'[^\d+]', '', phone)
            if len(phone) >= 10:
                return phone
    
    return ""

def extract_price(text: str) -> tuple:
    """Extrait le prix depuis un texte. Retourne (prix_formaté, prix_numérique)"""
    if not text:
        return ("", 0.0)
    
    # Patterns pour prix
    patterns = [
        r'(\d+[\s,.]?\d*)\s*€',  # 150 €, 150€, 1 500€
        r'(\d+[\s,.]?\d*)\s*euros?',  # 150 euros, 150 euro
        r'prix[:\s]+(\d+[\s,.]?\d*)',  # prix: 150
        r'tarif[:\s]+(\d+[\s,.]?\d*)',  # tarif: 150
        r'montant[:\s]+(\d+[\s,.]?\d*)',  # montant: 150
        r'(\d+[\s,.]?\d*)\s*EUR',  # 150 EUR
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            price_str = matches[0].replace(' ', '').replace(',', '.')
            # Vérifier que c'est un prix raisonnable (pas une année par exemple)
            try:
                price_num = float(price_str)
                if 5 <= price_num <= 100000:  # Prix raisonnable pour une prestation
                    # Formater avec virgule (format français) : 150,50 €
                    price_formatted = f"{price_num:.2f}".replace('.', ',') + " €"
                    return (price_formatted, price_num)
            except:
                pass
    
    return ("", 0.0)

def extract_client_name(text: str) -> str:
    """Extrait le nom du client depuis un texte"""
    if not text:
        return ""
    
    patterns = [
        r'client[:\s]+([^\n\r]+)',
        r'nom[:\s]+([^\n\r]+)',
        r'client\s*:?\s*([A-Z][a-zA-ZÀ-ÿ\s\-]+)',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            name = matches[0].strip()
            # Limiter la longueur
            if len(name) > 100:
                name = name[:100]
            return name
    
    return ""

def extract_company(text: str) -> str:
    """Extrait le nom de l'entreprise depuis un texte"""
    if not text:
        return ""
    
    patterns = [
        r'entreprise[:\s]+([^\n\r]+)',
        r'société[:\s]+([^\n\r]+)',
        r'company[:\s]+([^\n\r]+)',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            company = matches[0].strip()
            if len(company) > 100:
                company = company[:100]
            return company
    
    return ""

def extract_service_description(text: str) -> str:
    """Extrait la description de la prestation/service"""
    if not text:
        return ""
    
    # Chercher les lignes contenant "Service:", "Prestation:", etc.
    patterns = [
        r'service[:\s]+([^\n\r]+)',
        r'prestation[:\s]+([^\n\r]+)',
        r'description[:\s]+([^\n\r]+)',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            desc = matches[0].strip()
            if len(desc) > 200:
                desc = desc[:200]
            return desc
    
    # Si pas de pattern trouvé, retourner le texte nettoyé (sans téléphone et prix)
    cleaned = text
    # Retirer les numéros de téléphone
    cleaned = re.sub(r'(\+33\s?[1-9](?:[\s.-]?\d{2}){4})', '', cleaned)
    cleaned = re.sub(r'(0[1-9](?:[\s.-]?\d{2}){4})', '', cleaned)
    # Retirer les prix
    cleaned = re.sub(r'(\d+[\s,.]?\d*)\s*€', '', cleaned)
    cleaned = re.sub(r'(\d+[\s,.]?\d*)\s*euros?', '', cleaned, flags=re.IGNORECASE)
    # Nettoyer
    cleaned = ' '.join(cleaned.split())
    
    if len(cleaned) > 200:
        cleaned = cleaned[:200] + "..."
    
    return cleaned if cleaned else ""

def analyze_description(description: str, location: str = "") -> dict:
    """Analyse une description complète et extrait toutes les informations"""
    if not description:
        description = ""
    
    full_text = description
    if location:
        full_text = f"{full_text}\n{location}"
    
    prix_format, prix_num = extract_price(full_text)
    return {
        'telephone': extract_phone_number(full_text),
        'prix': prix_format,
        'prix_numerique': prix_num,
        'client': extract_client_name(description),
        'entreprise': extract_company(description),
        'prestation': extract_service_description(description),
        'description_brute': clean_description(description)
    }

def main():
    """Fonction principale pour exporter les prestations"""
    print("=" * 60)
    print("EXPORT DES PRESTATIONS GOOGLE CALENDAR")
    print("Période: 01/01/2025 - 08/01/2026")
    print("=" * 60)
    
    # Initialiser le service Google Calendar
    print("\n[INFO] Initialisation du service Google Calendar...")
    try:
        calendar_service = GoogleCalendarService()
        
        if not calendar_service.services:
            print("[ERREUR] Aucun calendrier configuré!")
            print("Vérifiez les fichiers credentials_adrien.json et credentials_clement.json")
            print("\nCONSEIL: Si les fichiers ont des problèmes de permissions, vous pouvez:")
            print("  1. Exécuter le script dans Docker: docker-compose exec backend python export_prestations_excel.py")
            print("  2. Vérifier que les variables d'environnement GOOGLE_SA_ADRIEN_JSON et GOOGLE_SA_CLEMENT_JSON sont définies")
            print("  3. Vérifier les permissions des fichiers credentials_*.json")
            return
        
        print(f"[OK] {len(calendar_service.services)} Service Account(s) configuré(s)")
        
        # Détecter tous les calendriers accessibles depuis chaque Service Account
        print("\n[INFO] Recherche de tous les calendriers accessibles...")
        
        # Initialiser avec les calendriers configurés pour ne pas les perdre
        all_calendars = dict(calendar_service.services)
        calendar_names_map = {}  # Map calendar_id -> nom
        
        # Pour chaque service configuré, lister tous les calendriers accessibles
        for calendar_id, service_info in calendar_service.services.items():
            service = service_info['service']
            base_name = service_info['name']
            
            try:
                # Lister tous les calendriers accessibles
                calendar_list = service.calendarList().list().execute()
                calendars = calendar_list.get('items', [])
                
                print(f"  [INFO] Service '{base_name}': {len(calendars)} calendrier(s) accessible(s)")
                
                for cal in calendars:
                    cal_id = cal.get('id')
                    cal_summary = cal.get('summary', cal_id)
                    cal_access_role = cal.get('accessRole', '')
                    
                    # Stocker le calendrier s'il n'est pas déjà connu ou s'il est partagé
                    if cal_id not in all_calendars:
                        # Déterminer le nom du calendrier
                        # Si c'est le calendrier principal, utiliser le nom du service
                        if cal_id == calendar_id:
                            cal_name = base_name
                        else:
                            # Pour les calendriers partagés ou secondaires, utiliser le summary
                            # Nettoyer le nom (enlever caractères spéciaux pour Excel)
                            cal_name = cal_summary.replace('/', '_').replace('\\', '_')[:31]
                        
                        all_calendars[cal_id] = {
                            'service': service,
                            'name': cal_name,
                            'summary': cal_summary,
                            'accessRole': cal_access_role
                        }
                        calendar_names_map[cal_id] = cal_name
                        print(f"    - Calendrier: {cal_summary} ({cal_id[:30]}...) -> Nom: {cal_name}")
                
            except HttpError as e:
                print(f"  [AVERTISSEMENT] Impossible de lister les calendriers pour '{base_name}': {e}")
                # Si on ne peut pas lister, garder au moins le calendrier principal configuré
                if calendar_id not in all_calendars:
                    all_calendars[calendar_id] = service_info
            except Exception as e:
                print(f"  [AVERTISSEMENT] Erreur pour '{base_name}': {e}")
                # Si on ne peut pas lister, garder au moins le calendrier principal configuré
                if calendar_id not in all_calendars:
                    all_calendars[calendar_id] = service_info
        
        # Vérifier s'il y a un troisième calendrier partagé configuré
        shared_calendar_id = (
            os.getenv('CALENDAR_ID_SHARED') 
            or os.getenv('CALENDAR_ID_COMMUN') 
            or os.getenv('CALENDAR_ID_PLANNING')
            or 'wash.go13@gmail.com'  # Calendrier commun par défaut
        )
        
        # Si un calendrier partagé est configuré, essayer de l'ajouter
        if shared_calendar_id and shared_calendar_id not in all_calendars:
            # Utiliser le premier service disponible pour tester l'accès au calendrier partagé
            if calendar_service.services:
                first_service_info = list(calendar_service.services.values())[0]
                service = first_service_info['service']
                
                # Tester si le calendrier est accessible
                try:
                    # Essayer de récupérer les métadonnées du calendrier pour vérifier l'accès
                    cal_info = service.calendars().get(calendarId=shared_calendar_id).execute()
                    cal_summary = cal_info.get('summary', 'Planning Commun')
                    
                    all_calendars[shared_calendar_id] = {
                        'service': service,
                        'name': 'Planning Commun',
                        'summary': cal_summary,
                        'accessRole': 'reader'
                    }
                    print(f"  [INFO] Calendrier partagé accessible: Planning Commun ({shared_calendar_id})")
                except HttpError as e:
                    if e.resp.status == 404:
                        print(f"  [AVERTISSEMENT] Calendrier '{shared_calendar_id}' non accessible. Assurez-vous qu'il est partagé avec les Service Accounts.")
                    else:
                        print(f"  [AVERTISSEMENT] Erreur d'accès au calendrier '{shared_calendar_id}': {e}")
                except Exception as e:
                    print(f"  [AVERTISSEMENT] Erreur lors de la vérification du calendrier '{shared_calendar_id}': {e}")
        
        # S'assurer qu'on a au moins les calendriers configurés
        if not all_calendars:
            print("  [INFO] Utilisation des calendriers configurés directement...")
            all_calendars = dict(calendar_service.services)
        
        # Mettre à jour le service avec tous les calendriers trouvés
        calendar_service.services = all_calendars
        calendars_list = ', '.join([info['name'] for info in all_calendars.values()])
        print(f"\n[OK] {len(all_calendars)} calendrier(s) trouvé(s) au total: {calendars_list}")
    except Exception as e:
        print(f"[ERREUR] Erreur lors de l'initialisation: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Définir la période
    date_debut = datetime(2025, 1, 1, 0, 0, 0)
    date_fin = datetime(2026, 1, 8, 23, 59, 59)
    
    print(f"\n[INFO] Récupération des événements...")
    print(f"   Du {format_date_fr(date_debut)} au {format_date_fr(date_fin)}")
    
    # Récupérer tous les événements
    try:
        events, warnings = calendar_service.get_events(
            calendar_ids=None,  # Tous les calendriers
            time_min=date_debut,
            time_max=date_fin,
            max_results=10000  # Augmenter pour être sûr de tout récupérer
        )
        
        if warnings:
            print("\n[AVERTISSEMENT]")
            for warning in warnings:
                print(f"   - {warning}")
        
        print(f"\n[OK] {len(events)} événement(s) récupéré(s)")
        
        if len(events) == 0:
            print("[ERREUR] Aucun événement trouvé dans la période spécifiée")
            return
        
    except Exception as e:
        print(f"[ERREUR] Erreur lors de la récupération des événements: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Préparer les données
    print("\n[INFO] Organisation des données...")
    prestations = []
    
    for event in events:
        start = event.get('start', {})
        end = event.get('end', {})
        
        start_str = start.get('dateTime') or start.get('date', '')
        end_str = end.get('dateTime') or end.get('date', '')
        
        start_dt = parse_datetime(start_str) if start_str else None
        end_dt = parse_datetime(end_str) if end_str else None
        
        # Déterminer si c'est un événement toute la journée
        is_all_day = 'date' in start and 'date' in end
        
        # Analyser la description pour extraire les informations
        description = event.get('description', '')
        location = event.get('location', '')
        analyzed = analyze_description(description, location)
        
        prestation = {
            'date': start_dt,
            'date_str': format_date_fr(start_dt) if start_dt else "",
            'heure_debut': start_dt.strftime("%H:%M") if start_dt and not is_all_day else "Journée entière",
            'heure_fin': end_dt.strftime("%H:%M") if end_dt and not is_all_day else "",
            'duree': extract_duration(start_str, end_str) if not is_all_day else "Journée entière",
            'calendrier': event.get('calendarName', 'Non défini'),
            'titre': event.get('summary', 'Sans titre'),
            'client': analyzed['client'],
            'entreprise': analyzed['entreprise'],
            'telephone': analyzed['telephone'],
            'prestation': analyzed['prestation'],
            'prix': analyzed['prix'],
            'prix_numerique': analyzed['prix_numerique'],
            'lieu': location or analyzed.get('lieu', ''),
            'description_complete': analyzed['description_brute'],
            'statut': event.get('status', 'confirmed'),
            'cree_le': parse_datetime(event.get('created', '')) if event.get('created') else None,
            'cree_le_str': format_datetime_fr(parse_datetime(event.get('created', ''))) if event.get('created') else "",
            'modifie_le': parse_datetime(event.get('updated', '')) if event.get('updated') else None,
            'modifie_le_str': format_datetime_fr(parse_datetime(event.get('updated', ''))) if event.get('updated') else "",
            'lien': event.get('htmlLink', ''),
        }
        
        prestations.append(prestation)
    
    # Trier par date (plus ancien en premier)
    prestations.sort(key=lambda x: x['date'] if x['date'] else datetime.min)
    
    # Séparer les prestations par calendrier
    prestations_par_calendrier = {}
    for prestation in prestations:
        calendrier = prestation['calendrier']
        if calendrier not in prestations_par_calendrier:
            prestations_par_calendrier[calendrier] = []
        prestations_par_calendrier[calendrier].append(prestation)
    
    print(f"[INFO] Prestations par calendrier: {', '.join([f'{k}: {len(v)}' for k, v in prestations_par_calendrier.items()])}")
    
    # Créer le classeur Excel
    print(f"\n[INFO] Création du fichier Excel...")
    wb = Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut
    
    # Définir les colonnes (ordre optimisé, sans la colonne Calendrier puisqu'on sépare)
    columns = [
        'Date',
        'Heure début',
        'Heure fin',
        'Durée',
        'Titre',
        'Client',
        'Entreprise',
        'Téléphone',
        'Description Prestation',
        'Prix',
        'Lieu',
        'Description complète',
        'Statut',
        'Créé le',
        'Modifié le',
        'Lien'
    ]
    
    # Style de l'en-tête
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Style de la ligne de totaux
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border_bottom = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thick')
    )
    
    # Largeurs des colonnes
    column_widths = {
        'Date': 12,
        'Heure début': 12,
        'Heure fin': 12,
        'Durée': 10,
        'Titre': 30,
        'Client': 25,
        'Entreprise': 25,
        'Téléphone': 15,
        'Description Prestation': 40,
        'Prix': 12,
        'Lieu': 25,
        'Description complète': 50,
        'Statut': 12,
        'Créé le': 18,
        'Modifié le': 18,
        'Lien': 60
    }
    
    prix_col_index = columns.index('Prix') + 1
    
    # Créer une feuille pour chaque calendrier
    for calendrier, prestations_cal in prestations_par_calendrier.items():
        # Créer la feuille avec un nom propre (limité à 31 caractères pour Excel)
        sheet_name = calendrier.capitalize()[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Écrire les en-têtes
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Ajuster les largeurs
        for col_num, column_title in enumerate(columns, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = column_widths.get(column_title, 15)
        
        # Écrire les données
        data_alignment = Alignment(vertical="top", wrap_text=True)
        total_prix = 0.0
        
        for row_num, prestation in enumerate(prestations_cal, 2):
            col = 1
            ws.cell(row=row_num, column=col).value = prestation['date_str']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['heure_debut']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['heure_fin']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['duree']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['titre']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['client']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['entreprise']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['telephone']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['prestation']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['prix']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['lieu']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['description_complete']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['statut']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['cree_le_str']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['modifie_le_str']; col += 1
            ws.cell(row=row_num, column=col).value = prestation['lien']
            
            # Ajouter au total
            total_prix += prestation.get('prix_numerique', 0.0)
            
            # Appliquer l'alignement et les bordures à toutes les cellules de la ligne
            for col_num in range(1, len(columns) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # Activer le retour à la ligne pour les colonnes de description
        desc_prestation_col = columns.index('Description Prestation') + 1
        desc_complete_col = columns.index('Description complète') + 1
        for row in range(2, len(prestations_cal) + 2):
            ws.cell(row=row, column=desc_prestation_col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row, column=desc_complete_col).alignment = Alignment(vertical="top", wrap_text=True)
        
        # Ajouter la ligne de totaux
        total_row = len(prestations_cal) + 2
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=prix_col_index).value = f"{total_prix:.2f}".replace('.', ',') + " €"
        
        # Formater la ligne de totaux
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=total_row, column=col_num)
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = Alignment(horizontal="right" if col_num == prix_col_index else "left", vertical="center")
            cell.border = thick_border_bottom if col_num <= prix_col_index else thin_border
        
        # Figer la première ligne (en-têtes)
        ws.freeze_panes = 'A2'
        
        # Appliquer un filtre automatique (sans inclure la ligne de totaux)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(prestations_cal) + 1}"
        
        total_formatted = f"{total_prix:.2f}".replace('.', ',')
        print(f"  [OK] Feuille '{sheet_name}' créée avec {len(prestations_cal)} prestations (Total: {total_formatted} €)")
    
    # Générer le nom du fichier avec la date actuelle
    date_export = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"prestations_2025_{date_export}.xlsx"
    filepath = backend_dir / filename
    
    # Sauvegarder le fichier
    wb.save(filepath)
    
    print(f"[OK] Export réussi!")
    print(f"\n[RESUME]")
    print(f"   - Nombre de prestations: {len(prestations)}")
    print(f"   - Période: 01/01/2025 - 08/01/2026")
    print(f"   - Fichier créé: {filename}")
    print(f"   - Chemin complet: {filepath.absolute()}")
    
    # Statistiques supplémentaires
    calendar_counts = {}
    for prestation in prestations:
        calendar = prestation['calendrier']
        calendar_counts[calendar] = calendar_counts.get(calendar, 0) + 1
    
    if calendar_counts:
        print(f"\n[STATISTIQUES] Par calendrier:")
        for calendar, count in sorted(calendar_counts.items()):
            print(f"   - {calendar}: {count} prestation(s)")

if __name__ == "__main__":
    main()
